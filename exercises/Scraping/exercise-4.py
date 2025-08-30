import openpyxl
from openpyxl import Workbook, load_workbook




def main():
    wb = openpyxl.load_workbook("marking-sample.xlsx",data_only=True)

    sheet = wb.active

    for row in sheet.iter_rows(min_row=4, max_row=7, values_only=True):
        print(row)

if __name__ == '__main__':
    main()
